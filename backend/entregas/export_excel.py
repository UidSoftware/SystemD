import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse


def exportar_excel(entregas, empresa, data_inicio, data_fim):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Entregas'

    uid_blue = 'FF063BF8'
    header_font = Font(bold=True, color='FFFFFFFF', size=10)
    header_fill = PatternFill('solid', fgColor=uid_blue)
    thin = Side(style='thin', color='FFE2E8F0')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    empresa_nome = empresa.nome_empresa if empresa else 'Todas as empresas'
    ws['A1'] = f'Uid Software — Relatório de Entregas'
    ws['A1'].font = Font(bold=True, size=13, color='FF063BF8')
    ws['A2'] = f'Empresa: {empresa_nome}'
    ws['A3'] = f'Período: {data_inicio or "—"} a {data_fim or "—"}'
    ws['A3'].font = Font(color='FF6b6b8a', italic=True)
    ws.append([])

    headers = ['Empresa', 'Data', 'Hora', 'Origem', 'Destino', 'Descrição', 'Status', 'Confirmação', 'Motivo']
    ws.append(headers)
    header_row = ws.max_row
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    alt_fill = PatternFill('solid', fgColor='FFF8F9FA')
    for i, e in enumerate(entregas):
        row = [
            e.empresa.nome_empresa,
            str(e.data),
            str(e.hora)[:5] if e.hora else '',
            e.origem,
            e.destino,
            e.descricao,
            e.get_status_display(),
            e.get_confirmacao_display(),
            e.confirmacao_motivo,
        ]
        ws.append(row)
        cur_row = ws.max_row
        for col_num in range(1, len(row) + 1):
            cell = ws.cell(row=cur_row, column=col_num)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if i % 2 == 1:
                cell.fill = alt_fill

    col_widths = [25, 12, 8, 30, 30, 35, 14, 16, 30]
    for col_num, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    nome = (empresa.nome_empresa.replace(' ', '_') if empresa else 'todas') + f'_{data_inicio}_{data_fim}'
    response['Content-Disposition'] = f'attachment; filename="entregas_{nome}.xlsx"'
    wb.save(response)
    return response
